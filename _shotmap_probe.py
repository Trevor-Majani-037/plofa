import json, sys, os
from openpyxl import load_workbook

xlsx = sys.argv[1]
try:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
except Exception as e:
    print("openpyxl error:", e)
    sys.exit(1)

print("SHEETS:", wb.sheetnames)

# --- Shot Map sheet ---
shot_sheet_rows = []
if "Shot Map" in wb.sheetnames:
    ws = wb["Shot Map"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    for r in rows[1:]:
        if r[0] is None:
            continue
        shot_sheet_rows.append(dict(zip(header, r)))
    print(f"\nSHOT MAP SHEET: {len(shot_sheet_rows)} shots")
    from collections import Counter
    teams = Counter(r["Team"] for r in shot_sheet_rows)
    outcomes = Counter(r["Outcome"] for r in shot_sheet_rows)
    print("per-team:", dict(teams))
    print("outcomes:", dict(outcomes))
    for r in shot_sheet_rows:
        print(f"  {r['Team'][:22]:22s} x={r['x']:6.1f} y={r['y']:6.1f} endX={r['End X']:6.1f} endY={r['End Y']:6.1f} {str(r['Outcome']):<9s} xG={r['xG']}")
else:
    print("No Shot Map sheet!")

# --- Team Summary ---
if "Team Summary" in wb.sheetnames:
    ws = wb["Team Summary"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    print("\nTEAM SUMMARY columns:", header)
    for r in rows[1:]:
        print(" ", dict(zip(header, r)))
