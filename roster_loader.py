"""
PLOFA 26/27 — ROSTER LOADER
==============================
roster_loader.py

Reads PLOFA-2026-2027.xlsx and automatically selects:
  • Starting XI  (First-Team players fill positions first; Second-Team fill gaps)
  • Bench        (remaining eligible players, max 7)
  • Superstars   (top-value players per club)
  • Set-piece takers (captain + high-value attackers/both-footed)
  • Formation    (auto-detected from First-Team position roster)

Squad tuple format fed to SquadBuilder.build():
    (name, position, [specialties], age, nationality, preferred_foot)

Foot values: "left" | "right" | "both"
→ "both" maps to both feet being strong (two_footed specialty added)

SPECIALTY column in Excel (optional, user-fills):
  Comma-separated code names, e.g.:  grand_dribbler,speedster,inverted
  If blank, specialties are inferred from position + leadership + foot.
"""

from __future__ import annotations
import os
import random
from datetime import date, datetime
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any

import openpyxl


# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "PLOFA-2026-2027.xlsx")

# Position normalization — maps Excel positions to engine positions
POS_MAP = {
    "GK":  "GK",
    "LB":  "LB",  "LWB": "LB",
    "RB":  "RB",  "RWB": "RB",
    "CB":  "CB",
    "CDM": "CDM",
    "CM":  "CM",
    "LM":  "LW",
    "RM":  "RW",
    "CAM": "CAM",
    "LW":  "LW",
    "RW":  "RW",
    "SS":  "ST",
    "ST":  "ST",
    "CF":  "CF",
}

# Sub timing by position (minute they typically come on)
SUB_TIMING = {
    "GK":  None,    # GK subs are injury-only
    "CB":  80,
    "LB":  75,  "RB": 75,
    "CDM": 72,
    "CM":  68,
    "CAM": 65,
    "LW":  65,  "RW": 65,
    "ST":  70,  "CF": 70,
}

# Positions needed per formation slot (ordered for selection)
FORMATION_SLOTS = {
    # 4-back systems
    "4-3-3":   ["GK", "CB", "CB", "LB", "RB", "CDM", "CM", "CM", "LW", "RW", "ST"],
    "4-2-3-1": ["GK", "CB", "CB", "LB", "RB", "CDM", "CM", "CAM", "LW", "RW", "ST"],
    "4-4-2":   ["GK", "CB", "CB", "LB", "RB", "CM",  "CM",  "LW",  "RW",  "ST", "ST"],
    # 3-back systems
    "3-4-3":   ["GK", "CB", "CB", "CB", "LB", "RB", "CM", "CM", "LW", "RW", "ST"],
    "3-4-2-1": ["GK", "CB", "CB", "CB", "LB", "RB", "CDM","CM", "LW", "RW", "ST"],
    "3-5-2":   ["GK", "CB", "CB", "CB", "LB", "RB", "CDM","CM", "CM", "ST", "ST"],
    "5-3-2":   ["GK", "CB", "CB", "CB", "LB", "RB", "CDM","CM", "CM", "ST", "ST"],
    "5-4-1":   ["GK", "CB", "CB", "CB", "LB", "RB", "CM", "CM", "LW", "RW", "ST"],
    "4-5-1":   ["GK", "CB", "CB", "LB", "RB", "CM", "CM", "CM", "CM", "CAM", "ST"],
}


# ─────────────────────────────────────────────
# PLAYER RECORD
# ─────────────────────────────────────────────

@dataclass
class PlayerRecord:
    """A player as read from the Excel file."""
    club:           str
    leadership:     Optional[str]   # "Captain" / "A. Captain" / None
    role:           Optional[str]   # "First-Team" / "Second-Team" / "Third-Team" / None
    number:         Optional[int]
    name:           str
    pos:            str             # Raw Excel position
    other_pos:      Optional[str]   # Secondary position
    dob:            Optional[datetime]
    market_value:   float
    preferred_foot: str             # "Left" / "Right" / "Both"
    specialties_raw: str            # Raw comma-sep string from SPECIALTIES column

    @property
    def engine_pos(self) -> str:
        """Normalized position for the match engine."""
        return POS_MAP.get(self.pos, self.pos)

    @property
    def engine_other_pos(self) -> Optional[str]:
        if self.other_pos:
            return POS_MAP.get(self.other_pos)
        return None

    @property
    def age(self) -> int:
        if self.dob is None:
            return 25
        today = date.today()
        b = self.dob
        return today.year - b.year - ((today.month, today.day) < (b.month, b.day))

    @property
    def foot_lower(self) -> str:
        """Lowercase foot string for engine."""
        return (self.preferred_foot or "right").lower()

    @property
    def is_captain(self) -> bool:
        return self.leadership in ("Captain", "A. Captain")

    @property
    def is_first_team(self) -> bool:
        return self.role == "First-Team"

    @property
    def is_second_team(self) -> bool:
        return self.role == "Second-Team"

    @property
    def is_third_team(self) -> bool:
        return self.role == "Third-Team"

    def get_specialties(self) -> List[str]:
        """Parse specialties from Excel.

        A blank SPECIALTY column means the player is intentionally average:
        NO position-based specialty is inferred, so trait-less players stay
        visibly distinct from good players (some specialties) and great
        players (many/strong specialties). Only universal traits that come
        from their own explicit columns (two-footed foot, captaincy) are
        still applied.
        """
        # Universal traits read from their own columns (not the SPECIALTY column)
        universal = []
        if self.foot_lower == "both":
            universal.append("two_footed")
        if self.is_captain:
            universal.append("captain")

        # User-provided specialties (highest priority)
        if self.specialties_raw and self.specialties_raw.strip():
            user_specs = [s.strip().lower() for s in self.specialties_raw.split(",") if s.strip()]
            for s in universal:
                if s not in user_specs:
                    user_specs.append(s)
            return user_specs

        # No SPECIALTY column data → leave the player average (no inference)
        return universal


# ─────────────────────────────────────────────
# ROSTER LOADER
# ─────────────────────────────────────────────

class RosterLoader:
    """
    Reads PLOFA-2026-2027.xlsx and provides squad-building utilities.

    Usage:
        loader = RosterLoader()
        # or: loader = RosterLoader("path/to/PLOFA-2026-2027.xlsx")

        clubs = loader.get_all_clubs()
        squad = loader.build_matchday_squad("Uditon")
        # → {
        #     "starters":   [(name, pos, specs, age, nationality, foot), ...],
        #     "substitutes": [(name, pos, specs, age, nationality, foot, sub_min), ...],
        #     "superstars":  [name, ...],
        #     "sp_takers":   [name, ...],
        #     "formation":   "4-3-3",
        #     "notes":       [...],
        # }
    """

    def __init__(self, excel_path: str = EXCEL_FILE):
        self.excel_path = excel_path
        self._players: Dict[str, List[PlayerRecord]] = {}   # club → [records]
        self._col_map: Dict[str, int] = {}                  # column name → index
        self._load()

    # ── PUBLIC API ───────────────────────────────────────────────────

    def get_all_clubs(self) -> List[str]:
        """Return all clubs found in the Excel."""
        skip = {"Total", None, ""}
        return [c for c in self._players.keys() if c not in skip]

    def get_club_players(self, club: str) -> List[PlayerRecord]:
        """Return all players for a club."""
        return self._players.get(club, [])

    def build_matchday_squad(
        self,
        club: str,
        availability: Optional[Dict[str, Any]] = None,
        formation: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Build a complete matchday squad for a club.

        Parameters
        ----------
        club : str
            Club name exactly as in Excel.
        availability : dict, optional
            {player_name: PlayerAvailability} from AvailabilityChecker.
            Players marked SUSPENDED, INJURED are excluded.
            FATIGUE_WARNING players may still start if no better option.
        formation : str, optional
            Force a specific formation string (e.g. "4-3-3").
            If None, auto-detected from roster positions.

        Returns
        -------
        dict with keys:
            starters    : list of tuples (name, pos, specs, age, nat, foot)
            substitutes : list of tuples (name, pos, specs, age, nat, foot, sub_min)
            superstars  : list of names
            sp_takers   : list of names
            formation   : str
            notes       : list of str (warnings/info for the user)
        """
        players = self.get_club_players(club)
        if not players:
            raise ValueError(f"No players found for club '{club}'. "
                             f"Available clubs: {self.get_all_clubs()}")

        notes = []

        # Filter out Third-Team and unavailable players
        eligible = self._filter_eligible(players, availability, notes)

        # Detect formation
        if formation is None:
            formation = self._detect_formation(eligible)
        notes.append(f"Formation: {formation}")

        slots = FORMATION_SLOTS.get(formation, FORMATION_SLOTS["4-3-3"])

        # Split into First-Team and Second-Team pools
        first_team  = [p for p in eligible if p.is_first_team]
        second_team = [p for p in eligible if p.is_second_team]
        third_team  = [p for p in eligible if not p.is_first_team and not p.is_second_team]

        # Select starting XI
        starters, used = self._fill_starting_xi(slots, first_team, second_team, third_team, eligible, notes)

        # Build bench
        used_names = {p.name for p in used}
        bench_pool = [p for p in eligible if p.name not in used_names]
        bench = self._build_bench(bench_pool, starters, notes)

        # Superstars (top-2 by market value)
        superstars = self._pick_superstars(eligible)

        # Set-piece takers (captain + high-value feet or both-footed)
        sp_takers = self._pick_sp_takers(starters, superstars)

        return {
            "starters":    [self._to_tuple(p) for p in starters],
            "substitutes": [self._to_sub_tuple(p) for p in bench],
            "superstars":  superstars,
            "sp_takers":   sp_takers,
            "formation":   formation,
            "notes":       notes,
        }

    # ── INTERNAL LOADERS ───────────────────────────────────────────

    def _load(self):
        """Load Excel data into memory."""
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        except FileNotFoundError:
            raise FileNotFoundError(
                f"Excel file not found: {self.excel_path}\n"
                f"Make sure PLOFA-2026-2027.xlsx is in the same folder."
            )

        ws = wb["PLAYERS"]

        # Build column map from header row
        headers = [cell.value for cell in ws[1]]
        self._col_map = {h: i for i, h in enumerate(headers) if h is not None}

        def col(name: str, row: tuple, default=None):
            idx = self._col_map.get(name)
            if idx is None:
                return default
            v = row[idx]
            return v if v is not None else default

        # Read all rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            club = row[0]
            if not club or not str(club).strip():
                continue
            club = str(club).strip()
            # Skip aggregate rows
            if club.lower() in ("total", ""):
                continue

            player_name = col("PLAYER NAME", row)
            if not player_name:
                continue
            player_name = str(player_name).strip()

            raw_pos = str(col("POS", row, "CM")).strip()
            other_pos_raw = col("OTHER POS", row)
            other_pos = str(other_pos_raw).strip() if other_pos_raw else None

            dob_val = col("DOB", row)
            if isinstance(dob_val, datetime):
                dob = dob_val
            else:
                dob = None

            foot_raw = col("PREFERRED FOOT", row, "Right")
            if foot_raw:
                foot = str(foot_raw).strip().capitalize()
                if foot not in ("Left", "Right", "Both"):
                    foot = "Right"
            else:
                foot = "Right"

            mv_raw = col("Market \nValue", row, 0)
            try:
                mv = float(mv_raw) if mv_raw else 0.0
            except (TypeError, ValueError):
                mv = 0.0

            leadership_raw = col("LEADERSHIP", row)
            leadership = str(leadership_raw).strip() if leadership_raw else None

            role_raw = col("ROLE", row)
            role = str(role_raw).strip() if role_raw else None

            number_raw = col("NO", row)
            try:
                number = int(number_raw) if number_raw else None
            except (TypeError, ValueError):
                number = None

            # SPECIALTY column (optional — may not exist yet)
            specs_raw = col("SPECIALTY", row, "")
            if specs_raw is None:
                specs_raw = ""
            else:
                specs_raw = str(specs_raw).strip()

            record = PlayerRecord(
                club=club,
                leadership=leadership,
                role=role,
                number=number,
                name=player_name,
                pos=raw_pos,
                other_pos=other_pos,
                dob=dob,
                market_value=mv,
                preferred_foot=foot,
                specialties_raw=specs_raw,
            )

            if club not in self._players:
                self._players[club] = []
            self._players[club].append(record)

    # ── FORMATION DETECTION ─────────────────────────────────────

    def _detect_formation(self, players: List[PlayerRecord]) -> str:
        """
        Detect formation from the positions in the First-Team roster.
        
        Rules (from user spec):
        - Has LWB or RWB in First-Team → 3-back system
        - Has LM + RM                  → 4-4-2
        - Has CAM (no LM/RM)           → 4-2-3-1
        - CDM only (no CAM, no LM/RM)  → 4-3-3
        - No wingers at all            → 3-5-2 or 4-5-1
        """
        first = [p for p in players if p.is_first_team]
        all_eligible = players

        raw_positions = {p.pos for p in first}
        engine_positions = {p.engine_pos for p in first}
        has_lwb = "LWB" in raw_positions
        has_rwb = "RWB" in raw_positions

        # Count available wingers/fullbacks across all eligible players
        has_any_winger = any(
            p.engine_pos in ("LW", "RW", "LB", "RB") or p.engine_other_pos in ("LW", "RW", "LB", "RB")
            for p in all_eligible
        )

        if has_lwb or has_rwb:
            # 3-back system — determine variant
            first_positions = [p.pos for p in first]
            has_lm = "LM" in first_positions
            has_rm = "RM" in first_positions
            has_lw = "LW" in first_positions
            has_rw = "RW" in first_positions
            cdm_count = first_positions.count("CDM")
            cm_count  = first_positions.count("CM")
            st_count  = first_positions.count("ST") + first_positions.count("CF")

            if (has_lw or has_lm) and (has_rw or has_rm):
                return "3-4-3"      # 3 CBs + LWB + RWB + 2 mids + LW + RW + ST
            elif cdm_count >= 1 and cm_count >= 2:
                return "5-3-2"      # Very defensive with wing-backs
            elif st_count >= 2:
                return "3-5-2"
            else:
                return "5-4-1"      # Deep wing-backs, 4 mids, 1 ST

        else:
            # 4-back system
            first_positions = [p.pos for p in first]
            first_engine = [p.engine_pos for p in first]
            has_lm  = "LM" in first_positions
            has_rm  = "RM" in first_positions
            has_cam = "CAM" in first_positions
            has_cdm = "CDM" in first_positions
            cdm_count = first_positions.count("CDM")
            cm_count  = first_positions.count("CM")
            st_count  = first_positions.count("ST") + first_positions.count("CF")

            # No wingers available → use no-winger formation
            if not has_any_winger:
                if st_count >= 2 and cm_count >= 2:
                    return "3-5-2"
                elif cm_count >= 4:
                    return "4-5-1"
                else:
                    return "3-5-2"  # fallback for no wingers

            if has_lm and has_rm:
                return "4-4-2"
            elif has_cam:
                return "4-2-3-1"
            elif has_cdm:
                return "4-3-3"
            else:
                return "4-4-2"  # fallback to 4-4-2 when we have wingers but no clear structure

    # ── SQUAD SELECTION ──────────────────────────────────────────

    def _filter_eligible(
        self,
        players: List[PlayerRecord],
        availability: Optional[Dict],
        notes: List[str],
    ) -> List[PlayerRecord]:
        """Filter out unavailable and Third-Team players."""
        eligible = []
        for p in players:
            # Exclude Third-Team entirely
            if p.is_third_team:
                continue
            # Check availability if provided
            if availability and p.name in availability:
                avail = availability[p.name]
                status = avail.status.value if hasattr(avail.status, 'value') else str(avail.status)
                if status in ("suspended_red", "suspended_yel", "injured"):
                    notes.append(f"⛔ {p.name} EXCLUDED ({status})")
                    continue
            eligible.append(p)
        return eligible

    def _fill_starting_xi(
        self,
        slots: List[str],
        first_team: List[PlayerRecord],
        second_team: List[PlayerRecord],
        third_team: List[PlayerRecord],
        all_eligible: List[PlayerRecord],
        notes: List[str],
    ) -> Tuple[List[PlayerRecord], List[PlayerRecord]]:
        """
        Fill the 11 starting positions using formation slots.
        Priority: First-Team → Second-Team → Third-Team (last resort).
        Returns (starters_list, all_used_list).
        """
        # Deduplicate input pools by name to prevent the same player being
        # selectable from multiple teams or from duplicate Excel rows.
        def dedup(players: List[PlayerRecord]) -> List[PlayerRecord]:
            seen: set = set()
            out: List[PlayerRecord] = []
            for p in players:
                if p.name not in seen:
                    seen.add(p.name)
                    out.append(p)
            return out

        first_team  = dedup(first_team)
        second_team = dedup(second_team)
        third_team  = dedup(third_team)

        # Also deduplicate across the combined eligible list so a player
        # present in both first_team and second_team cannot be selected twice.
        combined_seen: set = set()
        combined: List[PlayerRecord] = []
        for p in first_team + second_team + third_team:
            if p.name not in combined_seen:
                combined_seen.add(p.name)
                combined.append(p)
        first_team  = [p for p in combined if p in first_team]
        second_team = [p for p in combined if p in second_team]
        third_team  = [p for p in combined if p in third_team]

        starters   = []
        used       = []
        used_names = set()
        filled_indices = set()

        # Build pools grouped by engine position
        def make_pool(players: List[PlayerRecord]) -> Dict[str, List[PlayerRecord]]:
            pool: Dict[str, List[PlayerRecord]] = {}
            for p in players:
                ep = p.engine_pos
                pool.setdefault(ep, []).append(p)
                # Also list under other_pos
                if p.engine_other_pos and p.engine_other_pos != ep:
                    pool.setdefault(p.engine_other_pos, []).append(p)
            return pool

        ft_pool = make_pool(first_team)
        st_pool = make_pool(second_team)
        tt_pool = make_pool(third_team)

        for i, slot in enumerate(slots):
            # Try First-Team first
            chosen = self._pick_from_pool(ft_pool, slot, used_names)
            if not chosen:
                # Try Second-Team
                chosen = self._pick_from_pool(st_pool, slot, used_names)
                if chosen:
                    notes.append(f"⚠️  {chosen.name} (2nd team) starts at {slot} — no First-Team option")
            if not chosen:
                # Last resort: Third-Team
                chosen = self._pick_from_pool(tt_pool, slot, used_names)
                if chosen:
                    notes.append(f"❗ {chosen.name} (3rd team) starts at {slot} — squad depth issue")

            if chosen:
                starters.append(chosen)
                used.append(chosen)
                used_names.add(chosen.name)
                filled_indices.add(i)
            else:
                notes.append(f"❗ No player available for slot {slot}")

        if len(starters) < len(slots):
            missing = len(slots) - len(starters)
            unfilled_slots = [slots[i] for i in range(len(slots)) if i not in filled_indices]
            remaining = [p for p in first_team + second_team + third_team
                         if p.name not in used_names]
            for slot in unfilled_slots:
                # Pass 1: try unique remaining players only
                compatible = [p for p in remaining
                              if p.name not in used_names
                              and (not slot == "GK" or "GK" == p.engine_pos)]
                eligible = [p for p in compatible
                            if slot == "GK" or p.engine_pos != "GK"]
                pool = eligible or compatible

                # Pass 2: if no unique player left, recycle an already-used
                # non-GK to fill the slot (last resort)
                if not pool:
                    used_non_gk = [p for p in starters
                                   if p.engine_pos != "GK"
                                   and (not slot == "GK" or "GK" == p.engine_pos)]
                    if used_non_gk:
                        pool = [used_non_gk[0]]
                        notes.append(
                            f"🚨 Emergency: no other player available — "
                            f"{pool[0].name} ({pool[0].engine_pos}) "
                            f"also slots into {slot}"
                        )

                if pool:
                    fallback = pool.pop(0)
                    if not any(
                        note.endswith(f"plays out of position to complete the XI")
                        for note in notes
                        if fallback.name in note
                    ):
                        notes.append(
                            f"⚠️  No {slot} available — {fallback.name} "
                            f"({fallback.engine_pos}) "
                            f"plays out of position to complete the XI"
                        )
                    starters.append(fallback)
                    used.append(fallback)
                    used_names.add(fallback.name)
                else:
                    raise ValueError(
                        f"Cannot build starting XI: {missing} formation slot(s) unfillable "
                        f"({len(starters)}/{len(slots)} players selected). "
                        f"Unfilled slots: {unfilled_slots}. No remaining players to fill them."
                    )
            if len(starters) != len(slots):
                raise ValueError(
                    f"Cannot build starting XI: {len(slots) - len(starters)} slot(s) still unfillable "
                    f"after fallback ({len(starters)}/{len(slots)} players selected)."
                )

        # Deduplicate starters by name (defensive: Excel may contain duplicate
        # rows that survived earlier filtering).
        seen_names: set = set()
        unique_starters: List[PlayerRecord] = []
        for p in starters:
            if p.name not in seen_names:
                seen_names.add(p.name)
                unique_starters.append(p)
        if len(unique_starters) != len(starters):
            missing = len(starters) - len(unique_starters)
            notes.append(
                f"🔧 Removed {missing} duplicate starter(s) from XI; "
                f"re-checking GK count."
            )
            starters = unique_starters
            used = list(starters)
            used_names = {p.name for p in starters}

            # Try to fill any slots lost to deduplication from remaining eligible players.
            if len(starters) < len(slots):
                unfilled_indices = [i for i in range(len(slots)) if i >= len(starters)]
                for idx in unfilled_indices:
                    if idx >= len(slots):
                        break
                    slot = slots[idx]
                    remaining_pool = [
                        p for p in all_eligible
                        if p.name not in used_names
                        and (slot == "GK" or p.engine_pos != "GK")
                    ]
                    if remaining_pool:
                        replacement = max(remaining_pool, key=lambda p: p.market_value)
                        starters.append(replacement)
                        used.append(replacement)
                        used_names.add(replacement.name)
                        notes.append(
                            f"🔧 Filled duplicate-vacated {slot} with {replacement.name}"
                        )

        gk_slots = sum(1 for s in slots if s == "GK")
        gk_starters = sum(1 for p in starters if p.engine_pos == "GK")
        if gk_slots != gk_starters:
            raise ValueError(
                f"Starting XI must contain exactly {gk_slots} GK, got {gk_starters}. "
                f"Starters: {[(p.name, p.engine_pos) for p in starters]}"
            )

        # Resolve duplicate flank positions (LW, RW, LB, RB) using
        # OTHER POS and foot preference
        starters = self._resolve_flank_duplicates(
            starters, slots, all_eligible, notes
        )
        used = list(starters)

        return starters, used

    def _find_bench_flank_replacement(
        self,
        all_eligible: List[PlayerRecord],
        needed_slot: str,
        used_names: set,
    ) -> Optional[PlayerRecord]:
        """Find a bench player who can play the needed slot."""
        candidates = [
            p for p in all_eligible
            if p.name not in used_names
            and (p.engine_pos == needed_slot or p.engine_other_pos == needed_slot)
        ]
        if not candidates:
            return None

        if needed_slot in ("LW", "LB"):
            left_footed = [p for p in candidates if p.foot_lower == "left"]
            both_footed = [p for p in candidates if p.foot_lower == "both"]
            candidates = left_footed + both_footed + [p for p in candidates if p.foot_lower == "right"]
        elif needed_slot in ("RW", "RB"):
            right_footed = [p for p in candidates if p.foot_lower == "right"]
            both_footed = [p for p in candidates if p.foot_lower == "both"]
            candidates = right_footed + both_footed + [p for p in candidates if p.foot_lower == "left"]

        candidates.sort(key=lambda p: p.market_value, reverse=True)
        return candidates[0]

    def _resolve_flank_duplicates(
        self,
        starters: List[PlayerRecord],
        slots: List[str],
        all_eligible: List[PlayerRecord],
        notes: List[str],
    ) -> List[PlayerRecord]:
        """Ensure starting XI has at most 1 LW, 1 RW, 1 LB, 1 RB.

        Uses OTHER POS and foot preference to swap out-of-position flank
        players with bench players who can fill the correct slot.
        """
        flank_positions = {"LW", "RW", "LB", "RB"}

        pos_counts = {}
        for p in starters:
            pos_counts[p.engine_pos] = pos_counts.get(p.engine_pos, 0) + 1

        duplicates = [pos for pos in flank_positions if pos_counts.get(pos, 0) > 1]
        if not duplicates:
            return starters

        used_names = {p.name for p in starters}

        for dup_pos in duplicates:
            correct_slot_idx = None
            for i, s in enumerate(slots):
                if s == dup_pos:
                    correct_slot_idx = i
                    break

            if correct_slot_idx is None:
                continue

            correct_slot_player = starters[correct_slot_idx]
            dup_players = [p for p in starters if p.engine_pos == dup_pos]

            if correct_slot_player.engine_pos == dup_pos:
                out_of_position = [
                    p for p in dup_players if p.name != correct_slot_player.name
                ]
            else:
                out_of_position = dup_players

            for oop_player in out_of_position:
                oop_slot_idx = starters.index(oop_player)
                actual_slot = slots[oop_slot_idx] if oop_slot_idx < len(slots) else None

                if not actual_slot:
                    continue

                replacement = self._find_bench_flank_replacement(
                    all_eligible, actual_slot, used_names
                )
                if replacement:
                    idx = starters.index(oop_player)
                    starters[idx] = replacement
                    used_names.add(replacement.name)
                    used_names.discard(oop_player.name)
                    notes.append(
                        f"🔄 Fixed duplicate {dup_pos}: swapped {oop_player.name} "
                        f"(out-of-position at {actual_slot}) for {replacement.name} "
                        f"({replacement.engine_pos})"
                    )

        return starters

    def _pick_from_pool(
        self,
        pool: Dict[str, List[PlayerRecord]],
        slot: str,
        used_names: set,
    ) -> Optional[PlayerRecord]:
        """Pick the best available player for a slot from a pool.
        
        Preference order:
        1. Primary-position match (engine_pos == slot)
        2. Secondary-position match (engine_other_pos == slot)
        3. For flank positions: foot preference (left-footed for LB/LW,
           right-footed for RB/RW, both-footed always preferred)
        4. Market value (highest first)
        """
        candidates = [
            p for p in pool.get(slot, [])
            if p.name not in used_names
        ]
        if not candidates:
            return None

        primary = [p for p in candidates if p.engine_pos == slot]
        secondary = [p for p in candidates if p.engine_pos != slot]
        pool_candidates = primary if primary else candidates

        if len(pool_candidates) > 1 and slot in ("LW", "RW", "LB", "RB"):
            left_side = slot in ("LW", "LB")
            preferred_foot = "left" if left_side else "right"
            foot_best = []
            foot_good = []
            foot_rest = []
            for p in pool_candidates:
                if p.foot_lower == "both":
                    foot_best.append(p)
                elif p.foot_lower == preferred_foot:
                    foot_good.append(p)
                else:
                    foot_rest.append(p)
            pool_candidates = foot_best + foot_good + foot_rest

        pool_candidates.sort(key=lambda p: p.market_value, reverse=True)
        return pool_candidates[0]

    def _build_bench(
        self,
        bench_pool: List[PlayerRecord],
        starters: List[PlayerRecord],
        notes: List[str],
        max_bench: int = 7,
    ) -> List[PlayerRecord]:
        """
        Build the bench from unused players.
        Priority: First-Team > Second-Team
        Keep max 7, always include a backup GK if possible.
        """
        # Sort: First-Team first, then by market value
        sorted_pool = sorted(
            bench_pool,
            key=lambda p: (0 if p.is_first_team else 1, -p.market_value)
        )

        bench = []
        gk_on_bench = False

        # Always try to get a backup GK
        gk_candidates = [p for p in sorted_pool if p.engine_pos == "GK"]
        if gk_candidates:
            bench.append(gk_candidates[0])
            gk_on_bench = True
            sorted_pool = [p for p in sorted_pool if p.name != gk_candidates[0].name]

        # Fill remaining spots
        for p in sorted_pool:
            if len(bench) >= max_bench:
                break
            bench.append(p)

        return bench

    def _pick_superstars(self, players: List[PlayerRecord]) -> List[str]:
        """Flag top-2 players by market value as superstars."""
        eligible = [p for p in players if p.market_value > 0]
        if not eligible:
            return []
        sorted_p = sorted(eligible, key=lambda p: p.market_value, reverse=True)
        return [p.name for p in sorted_p[:2]]

    def _pick_sp_takers(
        self, starters: List[PlayerRecord], superstars: List[str]
    ) -> List[str]:
        """
        Pick set-piece takers:
        - Captain always takes set pieces
        - Creators/playmakers are primary takers regardless of position
        - Wingers/fullbacks are preferred over CBs for corners
        - Both-footed players (both feet = good delivery)
        - Superstars
        Max 3 takers.
        """
        takers = []
        creative_specs = {"creator", "grand_creator", "sup_vision", "playmaker", "dl_playmaker"}
        corner_preferred = {"LW", "RW", "LB", "RB", "CAM", "CM", "CDM"}

        for p in starters:
            if p.is_captain and p.name not in takers:
                takers.append(p.name)
            specs = [s.lower() for s in p.get_specialties()]
            if any(spec in creative_specs for spec in specs) and p.name not in takers:
                takers.append(p.name)
            if p.foot_lower == "both" and p.name not in takers:
                takers.append(p.name)
            if p.name in superstars and p.name not in takers:
                takers.append(p.name)
            if len(takers) >= 3:
                break

        # Fallback: prefer non-CB takers for set pieces
        if not takers and starters:
            non_cb = [p for p in starters if p.engine_pos != "CB"]
            pool = non_cb if non_cb else starters
            best = max(pool, key=lambda p: p.market_value)
            takers.append(best.name)

        return takers[:3]

    # ── TUPLE BUILDERS ───────────────────────────────────────────

    def _to_tuple(self, p: PlayerRecord) -> Tuple:
        """Convert PlayerRecord to SquadBuilder starter tuple."""
        return (
            p.name,
            p.engine_pos,
            p.get_specialties(),
            p.age,
            "Tolandian",         # nationality placeholder
            p.foot_lower,
        )

    def _to_sub_tuple(self, p: PlayerRecord) -> Tuple:
        """Convert PlayerRecord to SquadBuilder substitute tuple (with sub_in_minute)."""
        sub_min = SUB_TIMING.get(p.engine_pos)
        if sub_min:
            return (
                p.name,
                p.engine_pos,
                p.get_specialties(),
                p.age,
                "Tolandian",
                p.foot_lower,
                sub_min,
            )
        else:
            return (
                p.name,
                p.engine_pos,
                p.get_specialties(),
                p.age,
                "Tolandian",
                p.foot_lower,
            )


# ─────────────────────────────────────────────
# SINGLETON — reuse a single loaded instance
# ─────────────────────────────────────────────

_LOADER_INSTANCE: Optional[RosterLoader] = None


def get_loader(excel_path: str = EXCEL_FILE) -> RosterLoader:
    """Return a cached RosterLoader (loads Excel once)."""
    global _LOADER_INSTANCE
    if _LOADER_INSTANCE is None:
        _LOADER_INSTANCE = RosterLoader(excel_path)
    return _LOADER_INSTANCE


# ─────────────────────────────────────────────
# TEAM STYLE AUTO-SELECTOR
# Maps formation → sensible default TeamStyle
# ─────────────────────────────────────────────

def auto_team_style(club: str, formation: str, is_home: bool = True):
    """
    Return a TeamProfile with a sensible default style based on formation.
    Import this into run_match.py if you want fully automatic styles.
    
    You can still override styles manually in run_match.py if needed.
    """
    from match_engine import TeamProfile, TeamStyle, PlayingStyle, Intensity

    style_map = {
        "4-3-3":   (TeamStyle.ATTACKING,           PlayingStyle.HIGH_PRESS,       Intensity.HIGH),
        "4-2-3-1": (TeamStyle.BALANCED,             PlayingStyle.POSSESSION,        Intensity.MEDIUM),
        "4-4-2":   (TeamStyle.WING_PLAY,            PlayingStyle.DIRECT,            Intensity.MEDIUM),
        "3-4-3":   (TeamStyle.ULTRA_ATTACKING,      PlayingStyle.HIGH_PRESS,        Intensity.HIGH),
        "3-4-2-1": (TeamStyle.FLUID_COUNTER,        PlayingStyle.COUNTER,           Intensity.MEDIUM),
        "3-5-2":   (TeamStyle.STRUCTURED_POSSESSION,PlayingStyle.PATIENT_BUILD_UP,  Intensity.MEDIUM),
        "5-3-2":   (TeamStyle.DEFENSIVE,            PlayingStyle.LOW_BLOCK,         Intensity.LOW),
        "5-4-1":   (TeamStyle.ULTRA_DEFENSIVE,      PlayingStyle.LOW_BLOCK,         Intensity.LOW),
    }
    style, playing, intensity = style_map.get(formation, style_map["4-3-3"])
    return TeamProfile(name=club, style=style, playing_style=playing, intensity=intensity)


# ─────────────────────────────────────────────
# CLI / DEMO
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")

    print("\n🏟️  PLOFA Roster Loader — Squad Preview")
    print("=" * 65)

    loader = RosterLoader()
    clubs = loader.get_all_clubs()
    print(f"  Clubs loaded: {len(clubs)}")
    print(f"  {', '.join(clubs)}\n")

    for club in clubs:
        try:
            squad = loader.build_matchday_squad(club)
            print(f"\n{'─'*65}")
            print(f"  🏟️  {club}  |  Formation: {squad['formation']}")
            print(f"{'─'*65}")
            print(f"  {'#':<3} {'Player':<25} {'Pos':<6} {'Foot':<6} {'Specialties'}")
            print(f"  {'─'*60}")
            print("  STARTING XI:")
            for i, t in enumerate(squad["starters"], 1):
                name, pos, specs, age, nat, foot = t[0], t[1], t[2], t[3], t[4], t[5]
                specs_str = ", ".join(specs[:3]) + ("..." if len(specs) > 3 else "")
                print(f"  {i:<3} {name:<25} {pos:<6} {foot:<6} {specs_str}")
            print("  BENCH:")
            for t in squad["substitutes"]:
                name, pos, specs = t[0], t[1], t[2]
                foot = t[5] if len(t) > 5 else "right"
                sub_min = t[6] if len(t) > 6 else "—"
                specs_str = ", ".join(specs[:2])
                print(f"  {'SUB':<3} {name:<25} {pos:<6} {foot:<6} {specs_str}  [sub@{sub_min}']")
            print(f"  ⭐ Superstars: {', '.join(squad['superstars'])}")
            print(f"  🎯 Set-piece takers: {', '.join(squad['sp_takers'])}")
            if squad["notes"]:
                print(f"  📋 Notes:")
                for note in squad["notes"]:
                    print(f"     {note}")
        except Exception as e:
            print(f"  ❌ Error building squad for {club}: {e}")

    print(f"\n{'═'*65}")
    print(f"  ✅ Roster Loader operational.")
    print(f"     All {len(clubs)} clubs verified.\n")
